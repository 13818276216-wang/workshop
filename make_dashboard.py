# -*- coding: utf-8 -*-
import sys, json, math
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from collections import defaultdict
from datetime import datetime

# ========== 读数据 ==========
path = r"E:\袋鼠先生\袋鼠先生MeOS\你的MeOS\01_我的资料\订单与交付\销售单明细账.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
col = {h: i for i, h in enumerate(headers)}

rows_data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    rows_data.append(row)
wb.close()

print(f"总数据行: {len(rows_data)}")

# ========== 客户分类 ==========
SNACK_CUSTOMERS = {
    '宜春鸣忙食品有限公司',
    '厦门赵一鸣商业管理有限公司',
    '长沙很忙零食食品有限公司',
    '四川零食很忙食品有限公司',
    '湖北零食很忙食品有限公司',
    '山东赵一鸣商业管理有限公司',
    '芜湖赵一鸣商业管理有限公司',
    '西安鸣忙供应链有限公司',
    '长沙晓忙食品有限公司',
}
LIGHTNING_CUSTOMER = '袋鼠先生（滨州）营销有限公司'

def get_category(cname):
    if cname == LIGHTNING_CUSTOMER:
        return 'lightning'
    if cname in SNACK_CUSTOMERS:
        return 'snack'
    return 'dealer'

# ========== 读合同台账获取省区经理 ==========
import subprocess, csv, io, re
LARK_CLI = r'C:\Users\Administrator\.workbuddy\binaries\node\cli-connector-packages\lark-cli.cmd'
try:
    result = subprocess.run([
        LARK_CLI, 'sheets', '+csv-get',
        '--url', 'https://daishuxiansheng.feishu.cn/wiki/PZgMwdwlGimD60kypFOcBVasn9c',
        '--sheet-id', 'M4qpLc',
        '--range', 'A1:N61'
    ], capture_output=True, text=True, encoding='utf-8')
    resp = json.loads(result.stdout)
    csv_text = resp['data']['annotated_csv']
    raw_lines = csv_text.strip().split('\n')
    csv_rows = []
    current = None
    for line in raw_lines:
        m = re.match(r'\[row=(\d+)\]\s(.*)', line)
        if m:
            if current is not None:
                csv_rows.append(current)
            current = m.group(2)
        else:
            if current is not None:
                current += '\n' + line
    if current is not None:
        csv_rows.append(current)
    reader = csv.reader(io.StringIO('\n'.join(csv_rows)))
    contract_headers = next(reader)
    contract_col = {h: i for i, h in enumerate(contract_headers)}
    
    manager_map = {}
    for row in reader:
        if len(row) < 12:
            continue
        cname = str(row[contract_col.get('合作方', 2)] or '').strip()
        manager = str(row[contract_col.get('负责人', 11)] or '').strip()
        if cname and manager and manager != 'None':
            manager_map[cname] = manager
    print(f"成功从飞书读取合同台账，匹配到 {len(manager_map)} 个负责人的映射。")
except Exception as e:
    print("读取飞书合同台账失败:", e)
    manager_map = {}

# ========== 列索引 ==========
cname_idx = col['客户名称']
channel_idx = col['销售渠道']
pname_idx = col['货品名称']
qty_idx = col['数量']
price_idx = col['单价']
amt_idx = col['金额']
profit_idx = col['毛利']
time_idx = 29  # AD列：产品发货时间
order_idx = col['订单编号']

# 先扫一遍时间戳
all_times = []
for r in rows_data:
    ts = str(r[time_idx] or '').strip()
    if ts:
        all_times.append(ts[:10])
all_days = sorted(set(all_times))
today_key_calc = all_days[-1] if all_days else ''

# 上月同日
today_parts = today_key_calc.split('-') if today_key_calc else []
prev_month_same_day = ''
if len(today_parts) == 3:
    y, m, d = int(today_parts[0]), int(today_parts[1]), today_parts[2]
    if m == 1:
        prev_month_same_day = f'{y-1}-12-{d}'
    else:
        prev_month_same_day = f'{y}-{m-1:02d}-{d}'

now = datetime.now()
system_month = now.strftime('%Y-%m')
latest_data_month = today_key_calc[:7] if today_key_calc else system_month
current_month_calc = latest_data_month if latest_data_month < system_month else system_month
print(f"数据最后日期: {today_key_calc}, 当月: {current_month_calc}, 上月同日: {prev_month_same_day}")

# ========== 四维度统计容器 ==========
# 全量
total_sales = 0
total_profit = 0
total_orders = set()
total_dealers = set()
monthly = defaultdict(lambda: {'sales': 0, 'profit': 0, 'orders': set()})
daily = defaultdict(lambda: {'sales': 0, 'orders': set()})

# 分类维度
cat_stats = defaultdict(lambda: {
    'year_sales': 0, 'year_profit': 0,
    'month_sales': 0, 'month_profit': 0,
    'today_sales': 0,
    'orders': set(), 'dealers': set()
})

# 经销商维度（省区经理）
dealers = {}
dealers_today = defaultdict(float)
dealers_month = defaultdict(float)
dealers_first_month = {}
managers = {}
managers_today = defaultdict(float)
managers_month = defaultdict(lambda: {'sales': 0, 'profit': 0})
managers_prev_month_same_day = defaultdict(float)
managers_active_clients = defaultdict(set)   # 30天内有订单的经销商
dealers_last_order_day = {}                   # 经销商最后一笔订单日期
dealers_order_days = defaultdict(set)           # 经销商所有下单日期（去重）
products = {}
products_month = defaultdict(float)
products_order_days = defaultdict(set)

# 客户名称修正映射
cname_fix_map = {
    '杭州品强电子商务有限公司': '杭州品强电子商务有限公司',
    '西安海盒鲜食品有限公司': '西安海盒鲜食品有限公司',
    '郑州锦门商贸有限公司': '郑州锦门商贸有限公司',
}
manual_manager = {
    '三明市双方贸易有限公司': '朱晓亮',
}

# ========== 处理所有行 ==========
for r in rows_data:
    cname = str(r[cname_idx] or '').strip()
    channel = str(r[channel_idx] or '').strip()
    pname = str(r[pname_idx] or '').strip()
    qty = float(r[qty_idx] or 0)
    price = float(r[price_idx] or 0)
    amt = float(r[amt_idx] or 0)
    profit = float(r[profit_idx] or 0) if r[profit_idx] is not None else 0
    order_id = str(r[order_idx] or '')
    time_str = str(r[time_idx] or '')

    if qty == 0 and amt == 0:
        continue

    pos_amt = amt
    cat = get_category(cname)

    # === 全量统计 ===
    total_sales += pos_amt
    total_profit += profit
    total_orders.add(order_id)
    total_dealers.add(channel)

    if time_str:
        try:
            month_key = time_str[:7]
            monthly[month_key]['sales'] += pos_amt
            monthly[month_key]['profit'] += profit
            monthly[month_key]['orders'].add(order_id)
            daily[time_str[:10]]['sales'] += pos_amt
            daily[time_str[:10]]['orders'].add(order_id)
        except:
            pass

    # === 分类统计（含全量）===
    for cat_key in (cat, 'all'):
        cat_stats[cat_key]['year_sales'] += pos_amt
        cat_stats[cat_key]['year_profit'] += profit
        cat_stats[cat_key]['orders'].add(order_id)
        cat_stats[cat_key]['dealers'].add(channel)
        if time_str:
            try:
                if time_str[:7] == current_month_calc:
                    cat_stats[cat_key]['month_sales'] += pos_amt
                    cat_stats[cat_key]['month_profit'] += profit
                if time_str[:10] == today_key_calc:
                    cat_stats[cat_key]['today_sales'] += pos_amt
            except:
                pass

    # === 经销商维度（仅dealer类） ===
    if cat != 'dealer':
        continue

    # 修正客户名称
    if cname == '黄辉15257127683':
        cname = cname_fix_map.get(channel, cname)
    manager = manager_map.get(cname, '未知')
    if manager == '未知' and channel in manual_manager:
        manager = manual_manager[channel]

    if channel not in dealers:
        dealers[channel] = {'sales': 0, 'profit': 0, 'orders': set(), 'qty': 0, 'manager': manager}
    dealers[channel]['sales'] += pos_amt
    dealers[channel]['profit'] += profit
    dealers[channel]['orders'].add(order_id)
    dealers[channel]['qty'] += qty

    if manager not in managers:
        managers[manager] = {'sales': 0, 'profit': 0, 'orders': set(), 'clients': set()}
    managers[manager]['sales'] += pos_amt
    managers[manager]['profit'] += profit
    managers[manager]['orders'].add(order_id)
    managers[manager]['clients'].add(channel)

    if pname not in products:
        products[pname] = {'sales': 0, 'qty': 0, 'profit': 0, 'prices': [], 'dealers': set()}
    products[pname]['sales'] += pos_amt
    products[pname]['qty'] += qty
    products[pname]['profit'] += profit
    if price > 0:
        products[pname]['prices'].append(price)
    products[pname]['dealers'].add(channel)

    if time_str:
        try:
            month_key = time_str[:7]
            day_key = time_str[:10]
            if day_key == today_key_calc:
                managers_today[manager] += pos_amt
                dealers_today[channel] += pos_amt
            if month_key == current_month_calc:
                managers_month[manager]['sales'] += pos_amt
                managers_month[manager]['profit'] += profit
                dealers_month[channel] += pos_amt
                products_month[pname] += pos_amt
            if prev_month_same_day and time_str[:10] <= prev_month_same_day and time_str[:7] == prev_month_same_day[:7]:
                managers_prev_month_same_day[manager] += pos_amt
            if channel not in dealers_first_month or month_key < dealers_first_month[channel]:
                dealers_first_month[channel] = month_key
            # 更新经销商最近一笔订单日期
            if channel not in dealers_last_order_day or day_key > dealers_last_order_day[channel]:
                dealers_last_order_day[channel] = day_key
            dealers_order_days[channel].add(day_key)
            products_order_days[pname].add(day_key)
        except:
            pass

print(f"省区经理数: {len(managers)}")

# ========== 排序 ==========
month_keys = sorted(monthly.keys())
day_keys = sorted(daily.keys())
recent_days = day_keys[-14:] if len(day_keys) > 14 else day_keys

dealer_ranking = sorted(dealers.items(), key=lambda x: x[1]['sales'], reverse=True)
product_ranking = sorted(products.items(), key=lambda x: x[1]['sales'], reverse=True)
manager_ranking = sorted(managers.items(), key=lambda x: x[1]['sales'], reverse=True)

# ========== 计算活跃家数（30天内有订单） ==========
from datetime import timedelta
if today_key_calc:
    today_dt = datetime.strptime(today_key_calc, '%Y-%m-%d')
    cutoff_dt = today_dt - timedelta(days=30)
    cutoff_key = cutoff_dt.strftime('%Y-%m-%d')
    for ch, last_day in dealers_last_order_day.items():
        if last_day >= cutoff_key:
            mgr = dealers.get(ch, {}).get('manager', '未知')
            if mgr != '未知':
                managers_active_clients[mgr].add(ch)

# ========== 计算经销商30天活跃天数 ==========
dealers_active_days = {}
if today_key_calc:
    for ch, days_set in dealers_order_days.items():
        active_count = sum(1 for d in days_set if d >= cutoff_key)
        dealers_active_days[ch] = active_count

# ========== 计算产品30天内下单次数（独立下单天数）==========
products_active_days_30 = {}
if today_key_calc:
    for pname, days_set in products_order_days.items():
        active_count = sum(1 for d in days_set if d >= cutoff_key)
        products_active_days_30[pname] = active_count

# ========== 计算月数（每个经理从最早经销商首单月 → 当前月跨度） ==========
manager_earliest_month = {}
for mgr_name, mdata in managers.items():
    earliest = None
    for ch in mdata['clients']:
        ch_first = dealers_first_month.get(ch)
        if ch_first and (earliest is None or ch_first < earliest):
            earliest = ch_first
    manager_earliest_month[mgr_name] = earliest

current_dt = datetime.strptime(current_month_calc, '%Y-%m')
def calc_months(earliest_month):
    if earliest_month:
        e_dt = datetime.strptime(earliest_month, '%Y-%m')
        return (current_dt.year - e_dt.year) * 12 + (current_dt.month - e_dt.month) + 1
    return 1

# ========== 本月全量 ==========
month_sales = round(monthly[current_month_calc]['sales'], 2) if current_month_calc in monthly else 0
month_profit = round(monthly[current_month_calc]['profit'], 2) if current_month_calc in monthly else 0
month_orders = len(monthly[current_month_calc]['orders']) if current_month_calc in monthly else 0

# 今日
today_sales = 0
today_orders = 0
today_key = ''
if len(day_keys) > 0:
    today_key = day_keys[-1]
    today_sales = daily[today_key]['sales']
    today_orders = len(daily[today_key]['orders'])

def resolve_product_manager(dealer_set):
    mgrs = set()
    for ch in dealer_set:
        mgr = dealers.get(ch, {}).get('manager', '未知')
        if mgr != '未知':
            mgrs.add(mgr)
    if len(mgrs) == 1:
        return list(mgrs)[0]
    if len(mgrs) > 1:
        return '多区'
    return '未知'

def make_overview(cat_key):
    """生成单个分类的总览数据"""
    s = cat_stats[cat_key]
    yr_sales = round(s['year_sales'], 2)
    yr_profit = round(s['year_profit'], 2)
    yr_rate = round(yr_profit / yr_sales * 100, 1) if yr_sales > 0 else 0
    mo_sales = round(s['month_sales'], 2)
    mo_profit = round(s['month_profit'], 2)
    mo_rate = round(mo_profit / mo_sales * 100, 1) if mo_sales > 0 else 0
    return {
        'year_sales': yr_sales,
        'year_profit': yr_profit,
        'year_rate': yr_rate,
        'month_sales': mo_sales,
        'month_profit': mo_profit,
        'month_rate': mo_rate,
        'today_sales': round(s['today_sales'], 2),
        'dealers': len(s['dealers']),
        'orders': len(s['orders']),
    }

# ========== 构建 JSON ==========
data = {
    'update_time': datetime.now().strftime('%Y-%m-%d %H:%M'),
    # 四维度总览
    'overview_all': make_overview('all'),
    'overview_snack': make_overview('snack'),
    'overview_lightning': make_overview('lightning'),
    'overview_dealer': make_overview('dealer'),
    # 保持向后兼容
    'total_sales': round(total_sales, 2),
    'total_profit': round(total_profit, 2),
    'total_profit_rate': round(total_profit/total_sales*100, 1) if total_sales > 0 else 0,
    'total_dealers': len(total_dealers),
    'total_orders': len(total_orders),
    'total_rows': len(rows_data),
    'current_month': current_month_calc,
    'month_sales': month_sales,
    'month_profit': month_profit,
    'month_orders': month_orders,
    'month_profit_rate': round(month_profit/month_sales*100, 1) if month_sales > 0 else 0,
    'today_key': today_key,
    'today_sales': round(today_sales, 2),
    'today_orders': today_orders,
    'months': month_keys,
    'monthly_data': [{
        'month': m,
        'sales': round(monthly[m]['sales'], 2),
        'profit': round(monthly[m]['profit'], 2),
        'orders': len(monthly[m]['orders'])
    } for m in month_keys],
    'daily_data': [{
        'day': d[5:],
        'sales': round(daily[d]['sales'], 2),
        'orders': len(daily[d]['orders'])
    } for d in recent_days],
    'managers': [{
        'name': m[0],
        'clients': len(m[1]['clients']),
        'active_clients': len(managers_active_clients.get(m[0], set())),
        'today_sales': round(managers_today.get(m[0], 0), 2),
        'month_sales': round(managers_month[m[0]]['sales'], 2),
        'prev_month_sales': round(managers_prev_month_same_day.get(m[0], 0), 2),
        'sales': round(m[1]['sales'], 2),
        'orders': len(m[1]['orders']),
        'data_months': calc_months(manager_earliest_month.get(m[0])),
    } for m in manager_ranking],
    'dealers': [{
        'name': d[0],
        'manager': d[1]['manager'],
        'sales': round(d[1]['sales'], 2),
        'profit': round(d[1]['profit'], 2),
        'orders': len(d[1]['orders']),
        'qty': int(d[1]['qty']),
        'profit_rate': round(d[1]['profit']/d[1]['sales']*100, 1) if d[1]['sales'] > 0 else 0,
        'today_sales': round(dealers_today.get(d[0], 0), 2),
        'month_sales': round(dealers_month.get(d[0], 0), 2),
        'active_days': dealers_active_days.get(d[0], 0),
        'first_month': dealers_first_month.get(d[0], '-')
    } for d in dealer_ranking],
    'products': [{
        'name': p[0][:30],
        'sales': round(p[1]['sales'], 2),
        'month_sales': round(products_month.get(p[0], 0), 2),
        'qty': int(p[1]['qty']),
        'profit': round(p[1]['profit'], 2),
        'avg_price': round(sum(p[1]['prices'])/len(p[1]['prices']), 2) if p[1]['prices'] else 0,
        'dealers': len(p[1]['dealers']),
        'profit_rate': round(p[1]['profit']/p[1]['sales']*100, 1) if p[1]['sales'] > 0 else 0,
        'active_days_30': products_active_days_30.get(p[0], 0),
        'manager': resolve_product_manager(p[1]['dealers'])
    } for p in product_ranking]
}

with open(r'E:\袋鼠先生\github-dashboard\dashboard_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("\n数据已导出到 dashboard_data.json")
print(f"月度: {month_keys}")
